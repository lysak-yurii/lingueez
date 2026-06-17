# Lingueez — a desktop app for studying vocabulary across languages.
# Copyright (C) 2024-2026 Yurii Lysak
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.
#
# Additional terms under AGPL-3.0 section 7 apply to this program; see the
# NOTICE file distributed with this source for details.
#
# SPDX-License-Identifier: AGPL-3.0-or-later

"""PDF / Excel / CSV / TXT exporters, GUI-free.

All functions take a list of row dicts (keys: ID, RowNumber, Status,
Language1, Word1, Language2, Word2, Source, created_at) plus settings,
and write straight to the chosen path. Raise on failure.
"""
import logging
import os
import sqlite3

import pandas as pd
from openpyxl.styles import PatternFill
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Image as RepImage
from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle

from app.config import get_bool, get_float, get_int

# Display headers in the same column order as the table
EXPORT_COLUMNS = ["ID", "RowNumber", "Status", "Language1", "Word1", "Language2", "Word2", "Source", "created_at"]
EXPORT_HEADERS = {
    "ID": "ID", "RowNumber": "№", "Status": "Status", "Language1": "Language",
    "Word1": "Word", "Language2": "Translation", "Word2": "Word",
    "Source": "Source", "created_at": "Created at",
}
# Friendly labels for the settings dialog column pickers
EXPORT_COLUMN_LABELS = {
    "ID": "ID (database id)", "RowNumber": "№ (row number)", "Status": "Status",
    "Language1": "Language 1", "Word1": "Word 1",
    "Language2": "Language 2 (translation)", "Word2": "Word 2 (translation)",
    "Source": "Source", "created_at": "Created at",
}
# Definition columns are fetched from the database on demand (all formats)
EXTRA_COLUMNS = ["Definition", "Definition2"]
EXTRA_LABELS = {"Definition": "Definition 1", "Definition2": "Definition 2"}
PDF_WIDTH_DEFAULTS = {"ID": 0.5, "RowNumber": 0.5, "Status": 0.8, "Language1": 1.0,
                      "Word1": 1.6, "Language2": 1.0, "Word2": 1.6, "Source": 1.0,
                      "created_at": 1.2, "Definition": 2.5, "Definition2": 2.5}
BUILTIN_FONTS = ["Helvetica", "Times-Roman", "Courier"]
# Bundled .ttf fonts live here (relative to the app's working dir, in both dev and
# frozen builds). The reportlab built-ins above are Latin-only, so the default below
# is a bundled Cyrillic-capable font — Helvetica mangles Ukrainian PDFs.
FONTS_DIR = os.path.join('assets', 'fonts')
DEFAULT_FONT = "NotoSans-Regular"


def list_font_names(font_folder=FONTS_DIR):
    """Names (filename sans extension) of every .ttf in the fonts folder."""
    if not os.path.isdir(font_folder):
        return []
    return sorted(os.path.splitext(f)[0] for f in os.listdir(font_folder)
                  if f.endswith('.ttf'))


def register_fonts(font_folder=FONTS_DIR):
    """Register every .ttf in the fonts folder with reportlab; returns their names."""
    names = []
    for font_name in list_font_names(font_folder):
        try:
            pdfmetrics.registerFont(TTFont(font_name, os.path.join(font_folder, font_name + '.ttf')))
            names.append(font_name)
        except Exception as exc:
            logging.error(f"Failed to register font {font_name}: {exc}")
    return names


def _fetch_definitions(rows, db_path):
    """Definitions per word id: {ID: (Definition, Definition2)}."""
    ids = []
    for row in rows:
        try:
            ids.append(int(row.get("ID")))
        except (TypeError, ValueError):
            continue
    if not ids:
        return {}
    with sqlite3.connect(db_path) as connection:
        marks = ",".join("?" * len(ids))
        cursor = connection.execute(
            f"SELECT id, Definition, Definition2 FROM words WHERE id IN ({marks})", ids)
        return {r[0]: (r[1] or "", r[2] or "") for r in cursor.fetchall()}


def _rows_to_table(rows, exclude_columns, db_path=None):
    exclude = {c.strip() for c in exclude_columns}
    included = [c for c in EXPORT_COLUMNS if c not in exclude]
    headers = [EXPORT_HEADERS[c] for c in included]
    data = [[("" if row.get(c) is None else row.get(c)) for c in included] for row in rows]
    extras = [c for c in EXTRA_COLUMNS if c not in exclude] if db_path else []
    if extras:
        definitions = _fetch_definitions(rows, db_path)
        for row, table_row in zip(rows, data):
            d1, d2 = definitions.get(row.get("ID"), ("", ""))
            values = {"Definition": d1, "Definition2": d2}
            table_row.extend(values[c] for c in extras)
        headers += [EXTRA_LABELS[c] for c in extras]
    return headers, data, included + extras


def export_to_excel_file(rows, file_path, settings, db_path='dictionary.db'):
    sheet_name = settings.get("sheet_name", "Sheet1")
    start_row = get_int(settings, "start_row", 0)
    start_column = get_int(settings, "start_column", 0)
    exclude_columns = settings.get("exclude_columns_excel", "ID,Source").split(',')
    alternate_row_color = settings.get("alternate_row_color", "#e0e0e0")
    auto_column_width = get_bool(settings, "auto_column_width", True)
    freeze_panes = get_bool(settings, "freeze_panes", False)

    headers, data, included = _rows_to_table(rows, exclude_columns, db_path)
    if not included:
        raise ValueError("No columns selected for export — enable at least one column "
                         "in Settings → Export → Excel / CSV.")
    df = pd.DataFrame(data, columns=headers)

    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, startrow=start_row,
                    startcol=start_column, index=False)
        worksheet = writer.sheets[sheet_name]

        if alternate_row_color:
            color = alternate_row_color.replace('#', '')
            for row_idx in range(start_row + 2, len(df) + start_row + 2):
                if (row_idx - start_row - 1) % 2 == 0:
                    for col_idx in range(start_column, len(headers) + start_column):
                        cell = worksheet.cell(row=row_idx, column=col_idx + 1)
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        if auto_column_width:
            for column in worksheet.columns:
                max_length = max((len(str(cell.value)) for cell in column if cell.value is not None), default=0)
                worksheet.column_dimensions[column[0].column_letter].width = max_length + 2

        if freeze_panes:
            worksheet.freeze_panes = worksheet['A2']


def export_to_csv_file(rows, file_path, settings, db_path='dictionary.db'):
    delimiter = settings.get("csv_delimiter", ",")
    exclude_columns = settings.get("exclude_columns_excel", "ID,Source").split(',')
    headers, data, included = _rows_to_table(rows, exclude_columns, db_path)
    if not included:
        raise ValueError("No columns selected for export — enable at least one column "
                         "in Settings → Export → Excel / CSV.")
    pd.DataFrame(data, columns=headers).to_csv(file_path, sep=delimiter, index=False)


def export_to_txt_file(rows, file_path, settings, db_path='dictionary.db'):
    delimiter_setting = settings.get("txt_delimiter", "\\t")
    delimiter = "\t" if delimiter_setting in ("\\t",) or delimiter_setting.lower() == "tab" else delimiter_setting
    exclude_columns = [c.strip() for c in settings.get("exclude_columns_txt", "ID,Source").split(',')]
    include_headers = get_bool(settings, "txt_include_headers", True)
    header_lines = settings.get("txt_header_lines", "#separator:tab\\n#html:true\\n").replace('\\n', '\n')

    _, data, included = _rows_to_table(rows, exclude_columns, db_path)
    if not included:
        raise ValueError("No columns selected for export — enable at least one column "
                         "in Settings → Export → TXT.")

    with open(file_path, 'w', encoding='utf-8') as txt_file:
        if include_headers:
            txt_file.write(header_lines)
            if not header_lines.endswith('\n'):
                txt_file.write('\n')
        for row in data:
            txt_file.write(delimiter.join(str(v) for v in row) + "\n")


def export_to_pdf_file(rows, file_path, settings, db_path='dictionary.db'):
    """Render rows into a styled PDF table, optionally with definitions.

    Returns a list of warning strings (empty on a clean export).
    """
    warnings = []
    left_margin = get_float(settings, "left_margin", 10)
    right_margin = get_float(settings, "right_margin", 10)
    top_margin = get_float(settings, "top_margin", 10)
    bottom_margin = get_float(settings, "bottom_margin", 10)
    page_size = settings.get("page_size", "Letter")
    font_name = settings.get("font_name", DEFAULT_FONT)
    font_size = get_float(settings, "font_size", 10)
    leading = get_float(settings, "leading", 12)
    alignment = settings.get("alignment", "CENTER")
    header_bg_color = settings.get("header_bg_color", "grey")
    bg_color = settings.get("bg_color", "beige")
    text_color = settings.get("text_color", "whitesmoke")
    grid_color = settings.get("grid_color", "black")
    bg_image = settings.get("bg_image", "")
    exclude_columns = [c.strip() for c in
                       settings.get("exclude_columns", "ID,Source,created_at,Definition,Definition2").split(',')]

    alignment_code = {"LEFT": 0, "CENTER": 1, "RIGHT": 2}.get(alignment, 1)
    pagesize = {"Letter": letter, "A4": A4}.get(page_size, letter)

    headers, data_rows, included = _rows_to_table(rows, exclude_columns, db_path)

    if not included:
        raise ValueError("No columns selected for PDF export — enable at least one column "
                         "in Settings → Export → PDF.")

    weights = [get_float(settings, f"pdf_col_width_{c}", PDF_WIDTH_DEFAULTS[c]) for c in included]
    if get_bool(settings, "pdf_auto_widths", True):
        usable = pagesize[0] - left_margin - right_margin
        col_widths = [usable * w / sum(weights) for w in weights]
    else:
        col_widths = [w * inch for w in weights]

    available_fonts = set(BUILTIN_FONTS) | set(pdfmetrics.getRegisteredFontNames())
    if font_name not in available_fonts:
        fallback = DEFAULT_FONT if DEFAULT_FONT in available_fonts else "Helvetica"
        warnings.append(f"Font '{font_name}' is not available; used {fallback} instead.")
        font_name = fallback

    styles = getSampleStyleSheet()
    custom_style = ParagraphStyle(name='CustomFontStyle', fontName=font_name,
                                  fontSize=font_size, leading=leading, alignment=alignment_code)
    styles.add(custom_style)

    data = [headers] + [[Paragraph(str(v), custom_style) for v in table_row]
                        for table_row in data_rows]

    pdf = SimpleDocTemplate(file_path, pagesize=pagesize, rightMargin=right_margin,
                            leftMargin=left_margin, topMargin=top_margin, bottomMargin=bottom_margin)
    elements = []

    if bg_image and os.path.isfile(bg_image):
        try:
            bg = RepImage(bg_image)
            page_width, page_height = pagesize
            scale = min((page_width - left_margin - right_margin) / bg.imageWidth,
                        (page_height - top_margin - bottom_margin) / bg.imageHeight)
            bg.drawWidth = bg.imageWidth * scale
            bg.drawHeight = bg.imageHeight * scale
            elements.append(bg)
        except Exception as exc:
            logging.error(f"Error loading background image: {exc}")

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), header_bg_color),
        ('TEXTCOLOR', (0, 0), (-1, 0), text_color),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), font_name),
        ('FONTNAME', (0, 1), (-1, -1), font_name),
        ('FONTSIZE', (0, 1), (-1, -1), font_size),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), bg_color),
        ('GRID', (0, 0), (-1, -1), 1, grid_color),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)
    pdf.build(elements)
    return warnings
